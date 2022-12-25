from openpyxl import Workbook
from openpyxl import load_workbook

####################################### Create sheet #####################################
wb = Workbook() # object instantiate
ws = wb.active  # create "Sheet" sheet
ws = wb.create_sheet("Sheet1")
ws = wb.create_sheet("Sheet2")
ws = wb["Sheet"]

print(wb.sheetnames)    # get all sheets

####################################### Playing data #####################################

# access 1 cell
c = ws["A4"]
ws["A4"] = 4
print(c)

d = ws.cell(row=4, column=2, value=10)
print(d)

# access many cell
cell_range = ws['A1':'C2']
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

for row in ws.iter_rows(min_row=1, max_row=2, max_col=3):
    for cell in row:
        print(cell)

for row in ws.values:
    for value in row:
        print(value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# data store
c.value = 'hello, world'

####################################### Save to file #####################################
wb.save("my_excel.xlsx")

####################################### Load workbook #####################################
wb = load_workbook("my_excel.xlsx")
print(wb.sheetnames)